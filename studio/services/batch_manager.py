import json
import uuid
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

BASE = Path(__file__).resolve().parent.parent
BATCH_DIR = BASE / "batches"
BATCH_DIR.mkdir(exist_ok=True)

MAX_BATCH_ROWS = 250
REQUIRED_HEADERS = {"test_case", "golden_name"}


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _normalize_value(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _batch_path(batch_id: str) -> Path:
    return BATCH_DIR / f"{batch_id}.json"


def parse_workbook_bytes(contents: bytes) -> list[dict]:
    workbook = load_workbook(BytesIO(contents), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The uploaded workbook is empty.")

    headers = [_normalize_header(cell) for cell in rows[0]]
    if not any(headers):
        raise ValueError("The first row must contain header names.")

    missing = REQUIRED_HEADERS - set(h for h in headers if h)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    tasks: list[dict] = []
    for row_index, row in enumerate(rows[1:], start=2):
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record = {}
        for header, value in zip(headers, row):
            if header:
                record[header] = _normalize_value(value)
        if not record.get("test_case") or not record.get("golden_name"):
            raise ValueError(f"Row {row_index} must include both test_case and golden_name.")
        record.setdefault("browser", "msedge")
        record["status"] = "pending"
        record["workflowStatus"] = "pending"
        record["result"] = {}
        record["rowIndex"] = row_index
        record["rowId"] = str(uuid.uuid4())[:8]
        tasks.append(record)

    if not tasks:
        raise ValueError("No valid batch rows were found in the workbook.")
    if len(tasks) > MAX_BATCH_ROWS:
        raise ValueError(f"Batch upload exceeds maximum supported rows ({MAX_BATCH_ROWS}).")

    return tasks


def create_batch_from_excel(source_file_name: str, rows: list[dict]) -> dict:
    batch_id = str(uuid.uuid4())[:8]
    now = datetime.utcnow().isoformat() + "Z"
    batch = {
        "id": batch_id,
        "name": f"Batch {batch_id}",
        "sourceFileName": source_file_name,
        "status": "pending",
        "createdAt": now,
        "updatedAt": now,
        "rowCount": len(rows),
        "rows": rows,
    }
    save_batch(batch)
    return batch


def save_batch(batch: dict) -> dict:
    batch["updatedAt"] = datetime.utcnow().isoformat() + "Z"
    path = _batch_path(batch["id"])
    path.write_text(json.dumps(batch, indent=2))
    return batch


def load_batch(batch_id: str) -> dict:
    path = _batch_path(batch_id)
    if not path.exists():
        raise FileNotFoundError(f"Batch {batch_id} not found.")
    return json.loads(path.read_text())


def load_batches() -> list[dict]:
    batches: list[dict] = []
    for path in sorted(BATCH_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True):
        try:
            batches.append(json.loads(path.read_text()))
        except Exception:
            pass
    return batches


def batch_to_workbook(batch: dict) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    headers = [
        "rowId",
        "rowIndex",
        "golden_id",
        "test_case",
        "golden_name",
        "browser",
        "status",
        "workflowStatus",
        "sourceFileName",
        "result_message",
        "result_conclusion",
        "result_run_url",
    ]
    sheet.append(headers)

    for row in batch.get("rows", []):
        result = row.get("result", {}) if isinstance(row.get("result"), dict) else {}
        sheet.append([
            row.get("rowId", ""),
            row.get("rowIndex", ""),
            row.get("golden_id", ""),
            row.get("test_case", ""),
            row.get("golden_name", ""),
            row.get("browser", ""),
            row.get("status", ""),
            row.get("workflowStatus", ""),
            batch.get("sourceFileName", ""),
            result.get("message", ""),
            result.get("conclusion", ""),
            result.get("run_url", ""),
        ])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()
